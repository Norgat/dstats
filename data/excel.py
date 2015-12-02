# -*- encoding: utf-8 -*-

# TODO: Вынести общие версии Player, Team и Tournament в отдельный пакет.

# DOC: https://openpyxl.readthedocs.org/en/default/
import openpyxl as excel
import openpyxl.styles as excel_styles


def get_sheet_type(sheet):
  if sheet["A1"].value == "type":
    return sheet["B1"].value
  else:
    return None

ColorToCodes = {
  'FFFFFFFF': 0, # не играл на предыдущем турнире
  'FF00B0F0': 1, # играли вместе на предыдущем турнире в команде 1
  'FF0070C0': 2, # играли вместе на предыдущем турнире в команде 2
  'FF92D050': 3  # играл на предыдущем турнире
}

# кешируем обратную мэпу для код фона -> цвет
CodesToColor = {}
for color in ColorToCodes.keys():
  CodesToColor[ ColorToCodes[color] ] = color


class Player:
  """Dota 2 Player"""
  def __init__(self, nickname, team, sheet, row_num, col_num):
    self.Nickname = nickname.lower()
    self.Team = team
    self.Sheet = sheet
    self.Row = row_num
    self.Column = col_num
    self.Cell = sheet.cell(row = self.Row, column = self.Column)

    fgColor = self.Cell.fill.fgColor.rgb
    if type(fgColor) is str:
      self.PrevCode = ColorToCodes.get(fgColor, 0)
    else:
      self.PrevCode = 0

  def __eq__(self, other):
    return self.Nickname == other.Nickname

  def __hash__(self):
    return hash(self.Nickname)

  def __repr__(self):
    return "< %s | %s >" % (self.Nickname, self.Team.Tag)

  def write(self, **kwargs):
    sheet = self.Sheet
    if "sheet" in kwargs:
      sheet = kwargs["sheet"]
    sheet.cell(row = self.Row, column = self.Column).value = self.Nickname

    # стиль ячейки не может быть изменён, только перезаписан
    fill_pattern = excel_styles.PatternFill(patternType = 'solid', fgColor = excel_styles.Color(CodesToColor[self.PrevCode]))
    sheet.cell(row = self.Row, column = self.Column).fill = fill_pattern


class Team:
  """Dota 2 team class."""
  def __init__(self, sheet, row_num):
    self.Tag = str(sheet.cell(row = row_num, column = 1).value)
    self.Sheet = sheet
    self.Row = row_num

    self.Players = set([])
    for i in xrange(2, 7):
      self.Players.add(Player(str(sheet.cell(row = row_num, column = i).value), self, sheet, row_num, i))

    self.Top = int( sheet.cell(row = row_num, column = 7).value )

  def __repr__(self):
    return self.Tag

  def write(self, **kwargs):
    sheet = self.Sheet

    if "sheet" in kwargs:
      sheet = kwargs["sheet"]

    sheet.cell(row = self.Row, column = 1).value = self.Tag
    sheet.cell(row = self.Row, column = 7).value = self.Top


    for player in self.Players:
      player.write(**kwargs)

    if "topdiff" in kwargs:
      if kwargs["topdiff"]:
        sheet.cell(row = self.Row, column = 9).value = self.TopDiff


class Tournament:
  """Dota 2 tournament class."""
  def __init__(self, sheet):
    self.Title = str(sheet["B2"].value)
    self.NumberOfTeams = int(sheet["D1"].value)
    self.EndDate = sheet["D2"].value
    self.Sheet = sheet

    self.Teams = {}
    self.Players = {}

    for i in xrange(4, 4 + self.NumberOfTeams):
      tmp_team = Team(sheet, i)
      self.Teams[tmp_team.Tag] = tmp_team

      for player in tmp_team.Players:
        self.Players[player.Nickname] = player

  def __repr__(self):
    return "Tournament: %s" % self.Title

  def write(self, **kwargs):
    sheet = self.Sheet
    if "sheet" in kwargs:
      sheet = kwargs["sheet"]

    # write type
    sheet["A1"] = "type"
    sheet["B1"] = "tournament"

    # write title
    sheet["A2"] = "title"
    sheet["B2"] = self.Title

    # write Number of teams
    sheet["C1"] = "Number of teams"
    sheet["D1"] = self.NumberOfTeams

    # write data column titles
    sheet["A3"], sheet["B3"], sheet["C3"] = "Team", "Player 1", "Player 2"
    sheet["D3"], sheet["E3"], sheet["F3"] = "Player 3", "Player 4", "Player 5"
    sheet["G3"] = "Top"

    if "topdiff" in kwargs:
      if kwargs["topdiff"]:
        sheet["I3"] = "Top change"

    for tag in self.Teams.keys():
      self.Teams[tag].write(**kwargs)